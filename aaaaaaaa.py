import numpy as np
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook



def giai_he_pttt():
    try:
        n = int(nhap_n.get())
        phuong_trinh = []
        for i in range(n):
            pt = []
            for j in range(n):
                pt.append(float(nhap_bien[i][j].get()))
            phuong_trinh.append(pt)

        he_so = np.array(phuong_trinh)
        hang_so = np.array([float(nhap_hang_so[i].get()) for i in range(n)])

        ket_qua = np.linalg.solve(he_so, hang_so)

        ket_qua_text.delete(1.0, tk.END)
        ket_qua_text.insert(tk.END, "Nghiệm của hệ:\n{}".format(ket_qua))
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))


def reset_fields():
    nhap_n.delete(0, tk.END)
    ket_qua_text.delete(1.0, tk.END)
    for khung_phuong_trinh in cua_so.winfo_children():
        if isinstance(khung_phuong_trinh, tk.Frame):
            khung_phuong_trinh.destroy()
    nhap_bien.clear()
    nhap_hang_so.clear()


def xoa_truong_phuong_trinh():
    for danh_sach_bien in nhap_bien:
        for nhap_bien_var in danh_sach_bien:
            nhap_bien_var.delete(0, tk.END)
    for nhap_hang_so_var in nhap_hang_so:
        nhap_hang_so_var.delete(0, tk.END)
    ket_qua_text.delete(1.0, tk.END)


def tao_cac_truong_phuong_trinh():
    try:
        ket_qua_text.delete(1.0, tk.END)
        for khung_phuong_trinh in cua_so.winfo_children():
            if isinstance(khung_phuong_trinh, tk.Frame):
                khung_phuong_trinh.destroy()
        nhap_bien.clear()
        nhap_hang_so.clear()

        n = int(nhap_n.get())
        nhap_bien.clear()
        nhap_hang_so.clear()

        for i in range(n):
            khung_phuong_trinh = tk.Frame(cua_so)
            khung_phuong_trinh.pack()

            danh_sach_bien = []
            for j in range(n):
                nhap_bien_var = tk.Entry(khung_phuong_trinh, width=5)
                nhap_bien_var.pack(side=tk.LEFT)
                danh_sach_bien.append(nhap_bien_var)

            nhap_bien.append(danh_sach_bien)

            nhan_hang_so = tk.Label(khung_phuong_trinh, text=" = ")
            nhan_hang_so.pack(side=tk.LEFT)

            nhap_hang_so_var = tk.Entry(khung_phuong_trinh, width=5)
            nhap_hang_so_var.pack(side=tk.LEFT)
            nhap_hang_so.append(nhap_hang_so_var)

    except Exception as e:
        messagebox.showerror("Lỗi", str(e))


def import_from_excel():
    try:
        file_path = "Book1.xlsx"  # Replace with the actual file path
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Clear existing fields
        reset_fields()

        # Get the number of equations
        n = int(nhap_n.get())

        for i in range(n):
            khung_phuong_trinh = tk.Frame(cua_so)
            khung_phuong_trinh.pack()

            danh_sach_bien = []
            for j in range(n):
                cell_value = worksheet.cell(row=i + 1, column=j + 1).value
                nhap_bien_var = tk.Entry(khung_phuong_trinh, width=5)
                nhap_bien_var.pack(side=tk.LEFT)
                nhap_bien_var.insert(tk.END, cell_value)
                danh_sach_bien.append(nhap_bien_var)

            nhap_bien.append(danh_sach_bien)

            nhan_hang_so = tk.Label(khung_phuong_trinh, text=" = ")
            nhan_hang_so.pack(side=tk.LEFT)

            cell_value = worksheet.cell(row=i + 1, column=n + 1).value
            nhap_hang_so_var = tk.Entry(khung_phuong_trinh, width=5)
            nhap_hang_so_var.pack(side=tk.LEFT)
            nhap_hang_so_var.insert(tk.END, cell_value)
            nhap_hang_so.append(nhap_hang_so_var)

    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
# Create the main tkinter window
cua_so = tk.Tk()
cua_so.title("Giải hệ phương trình tuyến tính")
cua_so.geometry("400x300")

# Create the input fields
nhap_n_label = tk.Label(cua_so, text="Nhập số phương trình:")
nhap_n_label.pack()

nhap_n = tk.Entry(cua_so)
nhap_n.pack()

tao_cac_truong_button = tk.Button(cua_so, text="Tạo các trường phương trình", command=tao_cac_truong_phuong_trinh)
tao_cac_truong_button.pack()

import_excel_button = tk.Button(cua_so, text="Import từ tệp Excel", command=import_from_excel)
import_excel_button.pack()

ket_qua_text = tk.Text(cua_so, height=4)
ket_qua_text.pack()

giai_button = tk.Button(cua_so, text="Giải", command=giai_he_pttt)
giai_button.pack()

reset_button = tk.Button(cua_so, text="Reset", command=reset_fields)
reset_button.pack()

xoa_truong_button = tk.Button(cua_so, text="Xóa trường phương trình", command=xoa_truong_phuong_trinh)
xoa_truong_button.pack()

nhap_bien = []
nhap_hang_so = []

# Run the tkinter event loop
cua_so.mainloop()